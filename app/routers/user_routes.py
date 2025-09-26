import os
import shutil
import threading
import uuid
from datetime import date, timedelta
from typing import List, Dict, Any

import razorpay
from openpyxl import Workbook
import requests
from fastapi import APIRouter, HTTPException, Depends, Request, Form, UploadFile, File, Query
from openpyxl.reader.excel import load_workbook
from pydantic import Field
from pymysql import IntegrityError
from sqlalchemy import and_, func, or_, desc, cast
from sqlalchemy.orm import Session
from starlette import status
from starlette.responses import JSONResponse

from app.crud.user import generate_sequential_user_id, paginate, get_user_public_response, build_user_response, \
    create_or_update_view_notification, send_notification_to_all_user, get_world_mobile_codes, check_and_log_chat
from app.models.models import *
from app.routers.authenticate import send_otp_sms
from app.schemas.authentication import UserResponse

from app.schemas.user_schemas import *

from app.utils.authenticate import get_current_user, set_password, generate_otp, send_otp, hash_password
from app.utils.database import get_db
import humanize

from app.utils.authenticate import verify_password

router = APIRouter()
token="bbb51ae4a5563b43a4088f906ff6868f"

UPLOAD_DIR = "static/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)
signup_store: Dict[str, Any] = {}
otp_store: Dict[str, Any] = {}
_signup_lock = threading.Lock()


# put these imports at top of your routers file (or near other imports)

allowed_statuses = ["active", "paid", "exclusive", "Active", "Paid"]

# make sure these are imported from your project
# from app.database import get_db
# from app.models import User, SignupSession
# helpers: set_password, generate_sequential_user_id, generate_otp, send_otp
# constant: UPLOAD_DIR\
import logging
logger = logging.getLogger("uvicorn.error")
OTP_EXPIRY_MINUTES = 2
RESEND_THROTTLE_SECONDS = 30  # simple throttle between resends
import datetime as dt

@router.post("/signup123")
async def signup123(
    # main signup fields
    email: str = Form(None),
    password: str = Form(None),
    name: str = Form(None),
    gender: str = Form(None),
    dob: date = Form(None),
    maritalStatus: str = Form(None),
    education: str = Form(None),
    occupation: str = Form(None),
    language: str = Form(None),
    height: str = Form(None),
    diet: str = Form(None),
    smoke: str = Form(None),
    drink: str = Form(None),
    city_name: str = Form(None),
    postal: str = Form(None),
    state: str = Form(None),
    country: str = Form(None),
    mobile: str = Form(None),
    mobilecode: str = Form(None),
    partnerExpectations: str = Form(None),
    bio: str = Form(None),
    status: str = Form("active"),
    memtype: str = Form("Free"),
    membershipExpiryDate: date = Form(None),

    # ✅ photo uploads
    photo1: UploadFile = File(None), # MODIFIED: Name changed for consistency from profile_pic
    photo2: UploadFile = File(None),

    # control
    session_id: str = Form(None),
    otp: str = Form(None),
    action: str = Form(None),

    db: Session = Depends(get_db)
):
    # ---------------------------
    # helpers
    # ---------------------------
    def parse_session_id(sid_str: str):
        if not sid_str:
            return None
        try:
            return int(sid_str.strip())
        except ValueError:
            raise HTTPException(status_code=400, detail="session_id must be an integer")

    action_norm = action.strip().lower() if action else None
    otp_norm = otp.strip() if otp else None
    session_id_norm = parse_session_id(session_id)

    # ---------------------------
    # Phase A: Create signup session (from Step 1)
    # ---------------------------
    if not session_id_norm and not otp_norm and not action_norm:
        # Validate required fields for the first step
        for field_name, v in (("email", email), ("password", password), ("name", name), ("gender", gender), ("dob", dob), ("mobile", mobile)):
            if v is None:
                raise HTTPException(status_code=422, detail=f"Missing required field for step 1: {field_name}")

        if db.query(User).filter(User.email == email or User.mobile== mobile).first():
            raise HTTPException(status_code=400, detail="Email or Mobile already registered")

        otp_val = generate_otp()
        now_utc = dt.datetime.utcnow()
        hashed_pw = set_password(password)

        # If a session for this email already exists, update it. Otherwise, create a new one.
        session = db.query(SignupSession).filter(SignupSession.email == email).first()
        if not session:
            session = SignupSession(email=email, created_at=now_utc)
            db.add(session)

        # Update session with data from Step 1
        session.password = hashed_pw
        session.name = name
        session.gender = gender
        session.dob = dob
        session.mobile = mobile
        session.mobilecode = mobilecode
        session.otp = otp_val
        session.status = "pending" # Reset status

        db.commit()
        db.refresh(session)

        try:
            # You might want to move OTP sending to a background task
            send_otp(mobilecode or "", mobile or "", otp_val)
            otp_sent_status = True
        except Exception as e:
            logger.exception("send_otp failed: %s", e)
            otp_sent_status = False # Inform client if OTP sending failed

        return {"phase": "request_otp", "session_id": session.id, "otp_sent": otp_sent_status}

    # ---------------------------
    # Phase B: Resend OTP
    # ---------------------------
    # ADDED: Logic to handle OTP resend requests
    if session_id_norm and action_norm == "resend":
        session = db.query(SignupSession).filter(SignupSession.id == session_id_norm).first()
        if not session:
            raise HTTPException(status_code=404, detail="Signup session not found")

        otp_val = generate_otp()
        session.otp = otp_val
        db.commit()

        try:
            send_otp(session.mobilecode or "", session.mobile or "", otp_val)
            return {"message": "OTP has been resent."}
        except Exception as e:
            logger.exception("resend_otp failed: %s", e)
            raise HTTPException(status_code=500, detail="Failed to resend OTP.")


    # ---------------------------
    # Phase C: Update session and Confirm OTP (from Step 5)
    # ---------------------------
    if session_id_norm and otp_norm:
        session = db.query(SignupSession).filter(SignupSession.id == session_id_norm).first()
        if not session:
            raise HTTPException(status_code=404, detail="Signup session not found")
        if session.otp != otp_norm:
            raise HTTPException(status_code=400, detail="Invalid OTP")

        # MODIFIED: Update the session with all the final data from the last step
        # This is the key fix: We now process the rest of the form data here.
        session.maritalStatus = maritalStatus
        session.education = education
        session.occupation = occupation
        session.language = language
        session.height = height
        session.diet = diet
        session.smoke = smoke
        session.drink = drink
        session.city_name = city_name
        session.postal = postal
        session.state = state
        session.country = country
        session.partnerExpectations = partnerExpectations
        session.bio = bio
        session.memtype = memtype
        session.membershipExpiryDate = membershipExpiryDate

        # MODIFIED: Handle photo uploads during the final step
        now_utc = dt.datetime.utcnow()
        tmp1, tmp2 = session.photo1, session.photo2 # Keep old photos if new ones aren't uploaded
        for idx, pic in enumerate([photo1, photo2], start=1):
            if pic and pic.filename:
                ts = int(now_utc.timestamp())
                tmp_filename = f"tmp_{session.id}_{ts}_{idx}_{pic.filename}"
                tmp_path = os.path.join(UPLOAD_DIR, tmp_filename)
                with open(tmp_path, "wb") as buffer:
                    shutil.copyfileobj(pic.file, buffer)
                if idx == 1:
                    tmp1 = tmp_filename
                else:
                    tmp2 = tmp_filename
        session.photo1 = tmp1
        session.photo2 = tmp2

        # Now, create the final user from the completed session data
        user_id = generate_sequential_user_id(db)
        today = date.today()
        calc_age = today.year - session.dob.year - ((today.month, today.day) < (session.dob.month, session.dob.day))

        final1, final2 = "", ""
        for idx, tmp in enumerate([session.photo1, session.photo2], start=1):
            if tmp:
                tmp_path = os.path.join(UPLOAD_DIR, tmp)
                new_name = f"{user_id}_{idx}_{tmp.split('_',3)[-1]}" if '_' in tmp else f"{user_id}_{idx}_{tmp}"
                dst_path = os.path.join(UPLOAD_DIR, new_name)
                try:
                    if os.path.exists(tmp_path):
                        os.rename(tmp_path, dst_path)
                        if idx == 1: final1 = new_name
                        else: final2 = new_name
                except Exception:
                    # Fallback in case of rename error
                    if idx == 1: final1 = tmp
                    else: final2 = tmp

        new_user = User(
            id=user_id,
            email=session.email,
            password=session.password,
            name=session.name,
            gender=session.gender,
            dob=session.dob,
            age=calc_age,
            maritalStatus=session.maritalStatus,
            education=session.education,
            occupation=session.occupation,
            language=session.language,
            height=session.height,
            diet=session.diet,
            smoke=session.smoke,
            drink=session.drink,
            city_name=session.city_name,
            postal=session.postal,
            state=session.state,
            country=session.country,
            mobile=session.mobile,
            mobilecode=session.mobilecode,
            partnerExpectations=session.partnerExpectations,
            bio=session.bio,
            status="active",
            memtype=session.memtype or "Free",
            membershipExpiryDate=session.membershipExpiryDate,
            photo1=final1,
            photo2=final2,
            photo1Approve=False,
            photo2Approve=False,
            mobileverify=True,
            is_signup_complete=True,
            signup_completed_at=dt.datetime.now(india_tz),
            created_at=dt.datetime.now(india_tz),
        )

        db.add(new_user)
        db.delete(session) # Clean up the session table
        db.commit()
        db.refresh(new_user)

        return {
            "phase": "confirm_otp",
            "message": "Signup complete",
            "user": { "id": new_user.id, "email": new_user.email, "name": new_user.name },
        }

    # If none of the conditions are met, it's a bad request
    raise HTTPException(status_code=400, detail="Invalid request. Please provide session_id and otp, or initial signup data.")



@router.post("/signup")
async def signup(
        email: str = Form(...),
        password: str = Form(...),  # In production, hash the password!
        name: str = Form(...),
        gender: str = Form(...),
        dob: date = Form(...),
        maritalStatus: str = Form(None),
        education: str = Form(None),
        occupation: str = Form(None),
        language: str = Form(None),
        height: str = Form(None),
        diet: str = Form(None),
        smoke: str = Form(None),
        drink: str = Form(None),
        city_name: str = Form(None),
        postal: str = Form(None),
        state: str = Form(None),
        country: str = Form(None),
        mobile: str = Form(None),
        mobilecode: str = Form(None),
        partnerExpectations: str = Form(None),
        bio: str = Form(None),
        status: str = Form("active"),
        memtype: str = Form("Free"),
        membershipExpiryDate: date = Form(None),
        # Accept an optional profile picture upload
        profile_pic: UploadFile = File(None),
        db: Session = Depends(get_db)
):
    # Check if the user already exists
    existing_user = db.query(User).filter(User.email == email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Generate the next user ID (e.g., UD1, UD2, ...)
    #user_id = get_next_user_id(db)
    #with user_id_lock:
    user_id =  generate_sequential_user_id(db)
    # Save the profile picture if provided
    picture_url = ""
    filename1=""
    if profile_pic:
        filename1=f"{user_id}_{profile_pic.filename}"
        file_path = os.path.join(UPLOAD_DIR, f"{user_id}_{profile_pic.filename}")
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(profile_pic.file, buffer)
        picture_url = file_path  # This could also be a URL if served statically

    # (Optional) Calculate age from dob if not provided
    today = date.today()
    calc_age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

    # Create the new user instance
    new_user = User(
        id=user_id,
        email=email,
        password=set_password(password),  # Remember: in production, store a hashed password!
        name=name,
        gender=gender,
        dob=dob,
        age=calc_age,
        maritalStatus=maritalStatus,
        education=education,
        occupation=occupation,
        language=language,
        height=height,
        diet=diet,
        smoke=smoke,
        drink=drink,
        city_name=city_name,
        postal=postal,
        state=state,
        country=country,
        mobile=mobile,
        mobilecode=mobilecode,
        partnerExpectations=partnerExpectations,
        bio=bio,
        status=status,
        memtype=memtype,
        membershipExpiryDate=membershipExpiryDate,
        photo1=filename1
        # Set default values for other fields or leave them as defined by the model defaults
        #lastSeen=datetime.utcnow()
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {"message": "User successfully created", "user": {"id": new_user.id, "email": new_user.email}}




def save_uploaded_file(file: UploadFile, user_id: str, suffix: str = ""):
    if not file:
        return None

    # Generate secure filename
    file_ext = os.path.splitext(file.filename)[1]
    filename = f"{user_id}_{suffix}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, filename)

    # Save file
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return filename



# Step 1: Register Basic + Send OTP
 # or wherever your hash function is

otp_store = {}

# Step 1: Register basic info and send OTP

@router.post("/signup/step1")
def signup_step(
    email: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),
    mobilecode: str = Form(...),
    mobile: str = Form(...),
    dob: str = Form(...),
    db: Session = Depends(get_db)
):
    if password != confirm_password:
        raise HTTPException(status_code=400, detail="Passwords do not match")

    if db.query(User).filter(User.email == email).first():
        raise HTTPException(status_code=400, detail="Email already registered")

    existing_user = db.query(User).filter(User.mobile == mobile).first()
    if existing_user:
        if existing_user.mobileverify:
            raise HTTPException(status_code=400, detail="Mobile already verified")
        else:
            raise HTTPException(status_code=400, detail="Mobile already registered but not verified")

    try:
        dob_date = datetime.strptime(dob, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid DOB format. Use YYYY-MM-DD")

    user_id = generate_sequential_user_id(db)


    new_user = User(
        id=user_id,
        email=email,
        password=set_password(password),
        mobile=mobile,
        mobilecode=mobilecode,
        dob=dob_date,
        verify_status=False,
        verify_email=False,
        mobileverify=False
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    # Save OTP
    otp = generate_otp()
    otp_store[f"{mobilecode}{mobile}"] = {"otp": otp, "user_id": new_user.id}

    success = send_otp(mobilecode, mobile, otp)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to send OTP")

    return {
        "message": "OTP sent successfully",
        "user_id": new_user.id,
        "via": "SMS" if mobilecode == "91" else "WhatsApp"
    }

# Step 2: Verify OTP
@router.post("/signup/step2")
def verify_otp(
    mobilecode: str = Form(...),
    mobile: str = Form(...),
    otp: str = Form(...),
    db: Session = Depends(get_db)
):
    user_key = f"{mobilecode}{mobile}"
    data = otp_store.get(user_key)

    if not data or data["otp"] != otp:
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")

    user = db.query(User).filter(User.id == data["user_id"]).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.verify_email = True
    user.verify_status = True
    user.mobileverify = True

    db.commit()
    del otp_store[user_key]

    return {"message": "OTP verified. Proceed to Step 3.", "user_id": user.id}


@router.post("/signup/resend-otp")
def resend_otp(
    mobilecode: str = Form(...),
    mobile: str = Form(...),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(
        User.mobilecode == mobilecode,
        User.mobile == mobile
    ).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # ✅ Removed mobileverify check — allow OTP resend even if verified

    otp = generate_otp()
    otp_store[f"{mobilecode}{mobile}"] = {"otp": otp, "user_id": user.id}

    success = send_otp(mobilecode, mobile, otp)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to send OTP")

    return {
        "message": "OTP resent successfully",
        "user_id": user.id,
        "via": "SMS" if mobilecode == "91" else "WhatsApp"
    }



# Step 3: Profile Info
@router.post("/signup/step3")
def signup_step3(
    user_id: str = Form(...),
    name: str = Form(...),
    gender: str = Form(...),
    maritalStatus: str = Form(...),
    education: str = Form(...),
    occupation: str = Form(...),
    language: str = Form(...),
    height: str = Form(None),
    diet: str = Form(None),
    smoke: str = Form(None),
    drink: str = Form(None),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user or not user.mobileverify:
        raise HTTPException(status_code=404, detail="User not found or not verified")

    user.name = name
    user.gender = gender
    user.maritalStatus = maritalStatus
    user.education = education
    user.occupation = occupation
    user.language = language
    user.height = height
    user.diet = diet
    user.smoke = smoke
    user.drink = drink

    db.commit()
    return {"message": "Step 3 profile details saved"}

# Step 4: Final step - Location & Photos
from typing import Optional
from fastapi import Form

@router.post("/signup/step4")
async def signup_step4(
    user_id: str = Form(...),
    city_name: str = Form(...),
    postal: str = Form(None),
    state: str = Form(...),
    country: str = Form(...),
    partnerExpectations: str = Form(None),
    bio: str = Form(...),
    phonehide: Optional[bool] = Form(True),   # 👈 optional, defaults to False
    photo1: UploadFile = File(...),
    photo2: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.city_name = city_name
    user.postal = postal
    user.state = state
    user.country = country
    user.partnerExpectations = partnerExpectations
    user.bio = bio
    user.photo1 = save_uploaded_file(photo1, user.id, "photo1")
    user.photo2 = save_uploaded_file(photo2, user.id, "photo2") if photo2 else None
    user.photo1Approve = False
    user.photo2Approve = False
    user.is_signup_complete = True
    user.status = "active"
    user.memtype = "Free"

    # ✅ Set from request, default stays False
    user.phonehide = phonehide if phonehide is not None else False

    db.commit()
    return {"message": "Signup complete", "phonehide": user.phonehide}



#
# allowed_statuses = ["active", "paid", "exclusive", "Active", "Paid"]
# from fastapi.responses import JSONResponse

# @router.get("/get_users")
# def get_users(
#     gender: Optional[str] = Query(None),
#     status: Optional[str] = Query(None),
#     page: int = Query(1, ge=1),
#     limit: int = Query(10, ge=1, le=100),
#     db: Session = Depends(get_db),
#     current_user: User = Depends(get_current_user),
# ):
#     #allowed_statuses = ["active", "paid", "exclusive", "Active", "Paid"]
#
#     # ✅ Blocked users
#     if status == "blocked":
#         blocked_q = db.query(BlockedProfile).filter_by(blocker_id=current_user.id)
#         total, has_next, blocked = paginate(blocked_q, page, limit)
#         users = [b.blocked_user for b in blocked]
#         return {
#             "total": total,
#             "page": page,
#             "limit": limit,
#             "has_next": has_next,
#             "users": users
#         }
#
#     # ✅ Saved users
#     if status == "saveProfile":
#         saved_q = db.query(SavedProfile).filter_by(user_id=current_user.id)
#         total, has_next, saved = paginate(saved_q, page, limit)
#         users = [s.saved_user for s in saved]
#         return {
#             "total": total,
#             "page": page,
#             "limit": limit,
#             "has_next": has_next,
#             "users": users
#         }
#
#     # ✅ Main user query with allowed statuses
#
#     query = db.query(User).filter(
#         User.id != current_user.id,
#         User.status.in_(allowed_statuses)  # ✅ Filter only allowed statuses
#     )
#
#     # ✅ Gender filter
#     if gender and gender.lower() != "all":
#         query = query.filter(User.gender == gender)
#
#     # ✅ Online users filter
#     if status == "online":
#         query = query.filter(User.onlineUsers.is_(True))
#     if status in allowed_statuses:
#         query = db.query(User).filter(
#             User.id != current_user.id,
#             User.status.in_(allowed_statuses)  # ✅ Filter only allowed statuses
#         )
#     # ✅ Exclude blocked, matched, and saved users
#     blocked_ids = db.query(BlockedProfile.blocked_user_id).filter(
#         BlockedProfile.blocker_id == current_user.id
#     )
#     matched_ids = db.query(MatchRequest.receiver_id).filter(
#         MatchRequest.sender_id == current_user.id
#     )
#     saved_ids = db.query(SavedProfile.saved_user_id).filter(
#         SavedProfile.user_id == current_user.id
#     )
#
#     query = query.filter(
#         ~User.id.in_(blocked_ids),
#         ~User.id.in_(matched_ids),
#         ~User.id.in_(saved_ids)
#     )
#
#     # ✅ Pagination
#     total = query.count()
#     skip = (page - 1) * limit
#     users = query.offset(skip).limit(limit).all()
#     has_next = (page * limit) < total
#
#     return {
#         "total": total,
#         "page": page,
#         "limit": limit,
#         "has_next": has_next,
#         "users": users
#     }
#NEW SIGNUP

# place near top of file (after other globals)


# in-memory signup sessions (keyed by user_id)

_session_lock = threading.Lock()
signup_sessions: Dict[str, Dict[str, Any]] = {}

SIGNUP_TTL_SECONDS = 60 * 60  # 1 hour
OTP_TTL_SECONDS = 5 * 60  # 5 minutes


def _expires_at_ts(seconds: int) -> str:
    """Calculates an ISO 8601 expiration timestamp in UTC."""
    return (datetime.utcnow() + timedelta(seconds=seconds)).isoformat()


def save_signup_session(user_id: str, data: dict):
    """Saves a new signup session in a thread-safe manner."""
    with _session_lock:
        data_copy = dict(data)
        data_copy["_expires_at"] = _expires_at_ts(SIGNUP_TTL_SECONDS)
        signup_sessions[user_id] = data_copy


def load_signup_session(user_id: str) -> Optional[dict]:
    """Loads a signup session if it exists and is not expired."""
    with _session_lock:
        session = signup_sessions.get(user_id)
        if not session:
            return None

        # Check for session expiration
        exp = session.get("_expires_at")
        if not exp or datetime.utcnow() > datetime.fromisoformat(exp):
            signup_sessions.pop(user_id, None)  # Clean up expired session
            return None

        # Return a copy without internal keys
        copy = dict(session)
        copy.pop("_expires_at", None)
        copy.pop("_otp", None)
        copy.pop("_otp_expires_at", None)
        return copy


def del_signup_session(user_id: str):
    """Deletes a signup session."""
    with _session_lock:
        signup_sessions.pop(user_id, None)


def save_otp_for_temp_user(user_id: str, otp: str):
    """Saves an OTP to an existing session without race conditions."""
    with _session_lock:
        session = signup_sessions.get(user_id)
        # Only update if the session actually exists
        if session:
            session["_otp"] = otp
            session["_otp_expires_at"] = _expires_at_ts(OTP_TTL_SECONDS)
            # No need to write back, as we are modifying the dictionary's object directly


def load_otp_for_temp_user(user_id: str) -> Optional[str]:
    """Loads an OTP if the session and the OTP itself are not expired."""
    with _session_lock:
        session = signup_sessions.get(user_id)
        if not session:
            return None

        # Check for OTP expiration
        otp_exp = session.get("_otp_expires_at")
        if not otp_exp or datetime.utcnow() > datetime.fromisoformat(otp_exp):
            session.pop("_otp", None)  # Clean up expired OTP data
            session.pop("_otp_expires_at", None)
            return None

        return session.get("_otp")

# -----------------------------
# NEW SIGNUP Step 1 (no DB commit) - returns user_id
# -----------------------------
@router.post("/newsignup/step1")
def newsignup_step1(
    email: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),
    mobilecode: str = Form(...),
    mobile: str = Form(...),
    dob: str = Form(...),
    db: Session = Depends(get_db)
):
    if password != confirm_password:
        raise HTTPException(status_code=400, detail="Passwords do not match")

    # check against completed users only
    if db.query(User).filter(User.email == email, User.is_signup_complete == True).first():
        raise HTTPException(status_code=400, detail="Email already registered")

    if db.query(User).filter(User.mobile == mobile, User.is_signup_complete == True).first():
        raise HTTPException(status_code=400, detail="Mobile already registered")

    try:
        dob_date = datetime.strptime(dob, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid DOB format. Use YYYY-MM-DD")

    # generate sequential user_id but don’t commit to DB yet
    user_id = generate_sequential_user_id(db)

    # store session data in memory
    session_data = {
        "email": email,
        "password_hash": set_password(password),
        "mobilecode": mobilecode,
        "mobile": mobile,
        "dob": dob_date.isoformat(),
        "mobileverify": False,
        "profile": None,
        "created_at": datetime.utcnow()
    }
    save_signup_session(user_id, session_data)

    # generate OTP and store in session
    otp = generate_otp()
    session_data["_otp"] = otp
    session_data["_otp_expires_at"] = datetime.utcnow() + timedelta(minutes=5)  # 5 min expiry
    save_signup_session(user_id, session_data)

    success = send_otp(mobilecode, mobile, otp)
    if not success:
        del_signup_session(user_id)
        raise HTTPException(status_code=500, detail="Failed to send OTP")

    return {"message": "OTP sent successfully", "user_id": user_id, "via": "SMS" if mobilecode == "91" else "WhatsApp"}


# -----------------------------
# NEW SIGNUP Step 2 - verify OTP
# -----------------------------

OTP_GRACE_SECONDS = 0  # exact expiry

@router.post("/newsignup/step2")
def new_verify_otp(
    mobilecode: str = Form(...),
    mobile: str = Form(...),
    otp: str = Form(...),
    db: Session = Depends(get_db)
):
    norm_code = (mobilecode or "").lstrip("+").strip()
    norm_mobile = (mobile or "").lstrip("+").strip()
    otp_str = str(otp or "").strip()

    # 1) Check in-progress signup_sessions
    for uid, sess in list(signup_sessions.items()):
        sess_code = (sess.get("mobilecode") or "").lstrip("+").strip()
        sess_mobile = (sess.get("mobile") or "").lstrip("+").strip()
        if sess_code == norm_code and sess_mobile == norm_mobile:
            session_otp = sess.get("_otp")
            if not session_otp:
                raise HTTPException(status_code=400, detail="Invalid or expired OTP (session)")
            sess_expires = sess.get("_otp_expires_at")
            if sess_expires and isinstance(sess_expires, datetime):
                if datetime.utcnow() > sess_expires:
                    sess.pop("_otp", None)
                    sess.pop("_otp_expires_at", None)
                    signup_sessions[uid] = sess
                    raise HTTPException(status_code=400, detail="Invalid or expired OTP (session)")
            if str(session_otp).strip() != otp_str:
                raise HTTPException(status_code=400, detail="Invalid OTP (session)")
            sess["mobileverify"] = True
            sess.pop("_otp", None)
            sess.pop("_otp_expires_at", None)
            signup_sessions[uid] = sess
            return {"message": "OTP verified. Proceed to Step 3.", "user_id": uid}

    # 2) Fallback: otp_store check
    phone_key = f"{norm_code}{norm_mobile}"
    data = otp_store.get(phone_key) or otp_store.get(f"+{phone_key}") or otp_store.get(norm_mobile)
    if not data:
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")

    expires_at = data.get("expires_at")
    if expires_at and isinstance(expires_at, datetime):
        if datetime.utcnow() > (expires_at + timedelta(seconds=OTP_GRACE_SECONDS)):
            otp_store.pop(phone_key, None)
            raise HTTPException(status_code=400, detail="Invalid or expired OTP (store)")

    stored_otp = str(data.get("otp") or "").strip()
    if stored_otp != otp_str:
        raise HTTPException(status_code=400, detail="Invalid OTP")

    linked_user_id = data.get("user_id")
    if linked_user_id:
        user = db.query(User).filter(User.id == linked_user_id).first()
        if not user:
            otp_store.pop(phone_key, None)
            raise HTTPException(status_code=404, detail="User not found (stale OTP entry)")
        user.verify_email = True
        user.verify_status = True
        user.mobileverify = True
        db.commit()
        otp_store.pop(phone_key, None)
        return {"message": "OTP verified. Proceed to Step 3.", "user_id": user.id}

    data["claimed"] = True
    data["claimed_at"] = datetime.utcnow()
    otp_store[phone_key] = data
    return {"message": "OTP verified. Proceed to Step 3.", "user_id": None}


# -----------------------------
# NEW SIGNUP Resend OTP
# -----------------------------
OTP_TTL_SECONDS = 5 * 60

@router.post("/newsignup/resend-otp")
def resend_otp(
    mobilecode: str = Form(...),
    mobile: str = Form(...),
    db: Session = Depends(get_db)
):
    norm_code = (mobilecode or "").lstrip("+").strip()
    norm_mobile = (mobile or "").lstrip("+").strip()
    phone_key = f"{norm_code}{norm_mobile}"

    # throttle: 10s
    existing = otp_store.get(phone_key)
    if existing:
        last = existing.get("created_at")
        if last and (datetime.utcnow() - last).total_seconds() < 10:
            raise HTTPException(status_code=429, detail="Try again in a few seconds")

    user = db.query(User).filter(
        func.trim(User.mobilecode) == norm_code,
        func.trim(User.mobile) == norm_mobile
    ).first()

    otp = generate_otp()
    otp_store[phone_key] = {
        "otp": otp,
        "created_at": datetime.utcnow(),
        "expires_at": datetime.utcnow() + timedelta(seconds=OTP_TTL_SECONDS),
        "claimed": False,
        "user_id": user.id if user else None
    }

    success = send_otp(norm_code, norm_mobile, otp)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to send OTP")

    return {
        "message": "OTP resent successfully",
        "user_id": user.id if user else None,
        "via": "SMS" if norm_code == "91" else "WhatsApp"
    }

# -----------------------------
# NEW SIGNUP Step 3 - save profile info into session (not DB)
# -----------------------------
@router.post("/newsignup/step3")
def newsignup_step3(
    user_id: str = Form(...),
    name: str = Form(...),
    gender: str = Form(...),
    maritalStatus: str = Form(...),
    education: str = Form(...),
    occupation: str = Form(...),
    language: str = Form(...),
    height: str = Form(None),
    diet: str = Form(None),
    smoke: str = Form(None),
    drink: str = Form(None),
):
    data = load_signup_session(user_id)
    if not data:
        raise HTTPException(status_code=404, detail="Signup session not found or expired")

    if not data.get("mobileverify"):
        raise HTTPException(status_code=400, detail="Mobile not verified yet")

    data["profile"] = {
        "name": name,
        "gender": gender,
        "maritalStatus": maritalStatus,
        "education": education,
        "occupation": occupation,
        "language": language,
        "height": height,
        "diet": diet,
        "smoke": smoke,
        "drink": drink,
    }
    save_signup_session(user_id, data)

    return {"message": "Step 3 profile saved", "user_id": user_id}

# -----------------------------
# NEW SIGNUP Step 4 - finalize: create actual DB User
# -----------------------------
@router.post("/newsignup/step4")
async def newsignup_step4(
    user_id: str = Form(...),
    city_name: str = Form(...),
    postal: str = Form(None),
    state: str = Form(...),
    country: str = Form(...),
    partnerExpectations: str = Form(None),
    bio: str = Form(...),
    phonehide: Optional[bool] = Form(True),
    photo1: UploadFile = File(...),
    photo2: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    data = load_signup_session(user_id)
    if not data:
        raise HTTPException(status_code=404, detail="Signup session not found or expired")

    if not data.get("mobileverify"):
        raise HTTPException(status_code=400, detail="Mobile not verified")

    # Re-check uniqueness against completed users (prevent race)
    if db.query(User).filter(User.email == data["email"], User.is_signup_complete == True).first():
        raise HTTPException(status_code=400, detail="Email already registered")
    if db.query(User).filter(User.mobile == data["mobile"], User.is_signup_complete == True).first():
        raise HTTPException(status_code=400, detail="Mobile already registered")

    # Create final User
    new_user = User(
        id=user_id,
        email=data["email"],
        password=data["password_hash"],
        mobile=data["mobile"],
        mobilecode=data["mobilecode"],
        dob=data["dob"],
        verify_status=True,
        verify_email=False,
        mobileverify=True,
        is_signup_complete=True,
        status="active",
        memtype="Free",
        # profile fields from temp
        name=(data.get("profile") or {}).get("name"),
        gender=(data.get("profile") or {}).get("gender"),
        maritalStatus=(data.get("profile") or {}).get("maritalStatus"),
        education=(data.get("profile") or {}).get("education"),
        occupation=(data.get("profile") or {}).get("occupation"),
        language=(data.get("profile") or {}).get("language"),
        height=(data.get("profile") or {}).get("height"),
        diet=(data.get("profile") or {}).get("diet"),
        smoke=(data.get("profile") or {}).get("smoke"),
        drink=(data.get("profile") or {}).get("drink"),
        # step4-specific
        city_name=city_name,
        postal=postal,
        state=state,
        country=country,
        partnerExpectations=partnerExpectations,
        bio=bio,
        phonehide=phonehide if phonehide is not None else False,
    )

    try:
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Database error while creating user: {e}")

    # Save photos; if fail, delete created user to avoid partial row
    try:
        photo1_path = save_uploaded_file(photo1, new_user.id, "photo1")
        new_user.photo1 = photo1_path
    except Exception as e:
        db.delete(new_user)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Failed to save photo1: {e}")

    if photo2:
        try:
            photo2_path = save_uploaded_file(photo2, new_user.id, "photo2")
            new_user.photo2 = photo2_path
        except Exception as e:
            db.delete(new_user)
            db.commit()
            raise HTTPException(status_code=500, detail=f"Failed to save photo2: {e}")

    # finalize flags & commit
    new_user.photo1Approve = False
    new_user.photo2Approve = False
    new_user.verify_status = True
    new_user.verify_email = new_user.verify_email or False
    new_user.mobileverify = True

    db.commit()
    db.refresh(new_user)

    # cleanup session
    del_signup_session(user_id)

    # also cleanup any otp_store keyed by mobile (if present)
    key = f"{new_user.mobilecode}{new_user.mobile}"
    otp_store.pop(key, None)

    return {"message": "Signup complete", "user_id": new_user.id, "phonehide": new_user.phonehide}


@router.get("/get_users")
def get_users(
    gender: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Normalize status for case-insensitive matching
    status_lower = status.lower() if status else None

    allowed_status = ["active", "paid", "exclusive"]
    # Precompute excluded id subqueries
    blocked_ids = db.query(BlockedProfile.blocked_user_id).filter(
        BlockedProfile.blocker_id == current_user.id
    )
    matched_ids = db.query(MatchRequest.receiver_id).filter(
        MatchRequest.sender_id == current_user.id
    )
    saved_ids = db.query(SavedProfile.saved_user_id).filter(
        SavedProfile.user_id == current_user.id
    )

    # Special-case endpoints that return related models (blocked / saved)
    match status_lower:
        case "blocked":
            blocked_q = db.query(BlockedProfile).filter_by(blocker_id=current_user.id)
            total, has_next, blocked = paginate(blocked_q, page, limit)
            users = [b.blocked_user for b in blocked]
            return {
                "total": total,
                "page": page,
                "limit": limit,
                "has_next": has_next,
                "users": users
            }

        case "saveprofile":
            saved_q = db.query(SavedProfile).filter_by(user_id=current_user.id)
            total, has_next, saved = paginate(saved_q, page, limit)
            users = [s.saved_user for s in saved]
            return {
                "total": total,
                "page": page,
                "limit": limit,
                "has_next": has_next,
                "users": users
            }

        # For other statuses, build the base query once and then refine
        case _:
            # base query excludes current user
            query = db.query(User).filter(User.id != current_user.id)

            # If status asked is 'online'
            if status_lower == "online":
                query = query.filter(User.onlineUsers.is_(True))
            elif status_lower in allowed_status:
                # match the exact requested status (case-insensitive)
                query = query.filter(
                    User.id != current_user.id,
                    func.lower(User.status) == status_lower
                )

            # If status is one of allowed_statuses (case-insensitive)
            # elif status_lower in allowed_statuses:
            #     # use func.lower to compare DB value case-insensitively
            #     query = query.filter(func.lower(User.status).in_(list(allowed_statuses)))

            # If status is None or 'all' -> do not filter by status
            elif status_lower is None or status_lower == "all":
                query = query.filter(
                    User.id != current_user.id,
                    func.lower(User.status).in_(allowed_status)
                )
            else:
                # Unknown status -> return empty result (alternatively raise 400)

                return {
                    "total": 0,
                    "page": page,
                    "limit": limit,
                    "has_next": False,
                    "users": []
                }

    # Gender filter (apply after building status filters)
    if gender and gender.lower() != "all":
        query = query.filter(User.gender == gender,
                             func.lower(User.status).in_(allowed_status)
                             )

    # Exclude blocked, matched, and saved users
    query = query.filter(
        ~User.id.in_(blocked_ids),
        ~User.id.in_(matched_ids),
        ~User.id.in_(saved_ids),
        func.lower(User.status).in_(allowed_status)
    )

    # Pagination (count/offset/limit)
    total = query.count()
    skip = (page - 1) * limit
    users = query.offset(skip).limit(limit).all()
    has_next = (page * limit) < total

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "has_next": has_next,
        "users": users
    }


@router.get("/gender-counts")
def get_gender_counts(db: Session = Depends(get_db)):
    gender_counts = (
        db.query(User.gender, func.count(User.id))
        .group_by(User.gender)
        .all()
    )

    counts = {gender: count for gender, count in gender_counts}

    # Add online
    counts["online"] = db.query(User).filter(User.onlineUsers == True).count()


    counts["All"] = db.query(User).filter(User.status.in_(["active", "paid", "exclusive"])).count()

    # Add status-based
    for stat in ["active", "paid", "exclusive"]:
        counts[stat] = db.query(User).filter(User.status == stat).count()

    return JSONResponse(counts)


# 📦 Request body schema

@router.post("/requests/send")
def send_match_request(
    payload: RequestPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    receiver_id = payload.user_id

    # ✅ 1. Prevent self request
    if current_user.id == receiver_id:
        raise HTTPException(status_code=400, detail="Cannot send request to yourself.")

    # ✅ 2. Check receiver exists
    receiver = db.query(User).filter(User.id == receiver_id).first()
    if not receiver:
        raise HTTPException(status_code=404, detail="Receiver user not found.")

    # ✅ 3. Check if receiver blocked current_user or vice versa
    is_blocked = db.query(BlockedProfile).filter(
        ((BlockedProfile.blocker_id == current_user.id) & (BlockedProfile.blocked_user_id == receiver_id)) |
        ((BlockedProfile.blocker_id == receiver_id) & (BlockedProfile.blocked_user_id == current_user.id))
    ).first()

    if is_blocked:
        raise HTTPException(status_code=403, detail="You cannot send a request to this user.")

    # ✅ 4. Check for existing match request
    existing = db.query(MatchRequest).filter_by(
        sender_id=current_user.id,
        receiver_id=receiver_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Match request already sent.")

    # ✅ 5. Create new match request
    new_request = MatchRequest(
        sender_id=current_user.id,
        receiver_id=receiver_id,
        status="pending"
    )
    db.add(new_request)

    # ✅ 6. Add or update notification
    existing_notify = db.query(Notification).filter_by(
        sender_id=current_user.id,
        receiver_id=receiver_id,
        status=NotificationStatus.pending
    ).first()

    if existing_notify:
        existing_notify.created_at = datetime.now()
        existing_notify.is_read = False
        existing_notify.message = "You have a new match request"
    else:
        new_notify = Notification(
            sender_id=current_user.id,
            receiver_id=receiver_id,
            status=NotificationStatus.pending,
            message="You have a new match request",
            is_read=False
        )
        db.add(new_notify)

    db.commit()

    return {
        "success": True,
        "message": "Match request sent successfully."
    }

@router.post("/requests/respond")
def respond_to_match_request(
    payload: MatchRequestResponsePayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    sender_id = payload.user_id
    response_status = payload.status.lower()

    # ✅ 1. Validate status
    if response_status not in ["accepted", "rejected"]:
        raise HTTPException(status_code=400, detail="Invalid status. Use 'accepted' or 'rejected'.")

    # ✅ 2. Find the existing match request
    match_request = db.query(MatchRequest).filter_by(
        sender_id=sender_id,
        receiver_id=current_user.id
    ).first()

    if not match_request:
        raise HTTPException(status_code=404, detail="Match request not found.")

    # ✅ 3. Handle rejection
    if response_status == "rejected":
        db.delete(match_request)

        # ✅ Add or update notification
        existing_notify = db.query(Notification).filter_by(
            sender_id=sender_id,
            receiver_id=current_user.id,
            status=NotificationStatus.pending
        ).first()

        if existing_notify:
            existing_notify.created_at = datetime.now()
            existing_notify.is_read = False
            existing_notify.message = "Your match request was rejected."
            existing_notify.status = NotificationStatus.rejected
        else:
            new_notify = Notification(
                sender_id=sender_id,
                receiver_id=current_user.id,
                status=NotificationStatus.rejected,
                message="Your match request was rejected.",
                is_read=False
            )
            db.add(new_notify)

        db.commit()
        return {
            "success": True,
            "message": "Match request rejected and deleted."
        }

    # ✅ 4. Handle acceptance
    match_request.status = "accepted"

    existing_notify = db.query(Notification).filter_by(
        sender_id=sender_id,
        receiver_id=current_user.id,
        status=NotificationStatus.pending
    ).first()

    if existing_notify:
        existing_notify.created_at = datetime.now()
        existing_notify.is_read = False
        existing_notify.message = "Your match request was accepted."
        existing_notify.status = NotificationStatus.accepted
    else:
        new_notify = Notification(
            sender_id=sender_id,
            receiver_id=current_user.id,
            status=NotificationStatus.accepted,
            message="Your match request was accepted.",
            is_read=False
        )
        db.add(new_notify)

    db.commit()
    return {
        "success": True,
        "message": "Match request accepted."
    }



@router.post("/save-unsave")
def save_unsave_profile(
    payload: RequestPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    saved_user_id = payload.user_id

    if current_user.id == saved_user_id:
        raise HTTPException(status_code=400, detail="Cannot save your own profile.")

    user_exists = db.query(User).filter(User.id == saved_user_id).first()
    if not user_exists:
        raise HTTPException(status_code=404, detail="User not found.")

    existing = db.query(SavedProfile).filter_by(
        user_id=current_user.id,
        saved_user_id=saved_user_id
    ).first()

    if existing:
        db.delete(existing)
        db.commit()
        return {"success": False, "message": "Profile unsaved"}
    else:
        new_saved = SavedProfile(user_id=current_user.id, saved_user_id=saved_user_id)
        db.add(new_saved)
        db.commit()
        return {"success": True, "message": "Profile saved"}




@router.post("/block-unblock")
def block_unblock_user(
    payload: BlockUserPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    blocked_user_id = payload.user_id

    # ❌ Cannot block yourself
    if current_user.id == blocked_user_id:
        raise HTTPException(status_code=400, detail="Cannot block yourself.")

    # ✅ Check if target user exists
    user_exists = db.query(User).filter(User.id == blocked_user_id).first()
    if not user_exists:
        raise HTTPException(status_code=404, detail="User not found.")

    # ✅ Check if already blocked
    existing_block = db.query(BlockedProfile).filter_by(
        blocker_id=current_user.id,
        blocked_user_id=blocked_user_id
    ).first()

    if existing_block:
        db.delete(existing_block)
        db.commit()
        return {"success": False, "message": "User unblocked"}

    # ✅ If not blocked: proceed to block and clean up related data
    new_block = BlockedProfile(
        blocker_id=current_user.id,
        blocked_user_id=blocked_user_id,
        reason=payload.reason
    )
    db.add(new_block)

    # ✅ Remove any saved profiles in either direction
    db.query(SavedProfile).filter(
        ((SavedProfile.user_id == current_user.id) & (SavedProfile.saved_user_id == blocked_user_id)) |
        ((SavedProfile.user_id == blocked_user_id) & (SavedProfile.saved_user_id == current_user.id))
    ).delete(synchronize_session=False)

    # ✅ Remove match requests in either direction
    db.query(MatchRequest).filter(
        ((MatchRequest.sender_id == current_user.id) & (MatchRequest.receiver_id == blocked_user_id)) |
        ((MatchRequest.sender_id == blocked_user_id) & (MatchRequest.receiver_id == current_user.id))
    ).delete(synchronize_session=False)

    # ✅ Remove notifications in either direction
    db.query(Notification).filter(
        ((Notification.sender_id == current_user.id) & (Notification.receiver_id == blocked_user_id)) |
        ((Notification.sender_id == blocked_user_id) & (Notification.receiver_id == current_user.id))
    ).delete(synchronize_session=False)

    db.commit()
    return {"success": True, "message": "User blocked"}




@router.get("/users/blocked")
def get_blocked_users(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(BlockedProfile).filter_by(blocker_id=current_user.id)
    total, has_next, blocked = paginate(query, page, limit)
    users = [b.blocked_user for b in blocked]
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "has_next": has_next,
        "users": users
    }


@router.get("/users/saved")
def get_saved_profiles(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(SavedProfile).filter_by(user_id=current_user.id)
    total, has_next, saved = paginate(query, page, limit)
    users = [s.saved_user for s in saved]
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "has_next": has_next,
        "users": users
    }




@router.get("/connections/received")
def get_received_requests(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(MatchRequest).filter(
        MatchRequest.receiver_id == current_user.id,
        MatchRequest.status == "pending"
    ).join(User, MatchRequest.sender_id == User.id)

    total = query.count()
    skip = (page - 1) * limit
    items = query.offset(skip).limit(limit).all()

    users = [req.sender for req in items]
    has_next = skip + limit < total
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "has_next": has_next,
        "users": users
    }


@router.get("/connections/sent")
def get_sent_requests(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(MatchRequest).filter(
        MatchRequest.sender_id == current_user.id,
        MatchRequest.status == "pending"
    ).join(User, MatchRequest.receiver_id == User.id)

    total = query.count()
    skip = (page - 1) * limit
    items = query.offset(skip).limit(limit).all()

    users = [req.receiver for req in items]
    has_next = skip + limit < total
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "has_next": has_next,
        "users": users
    }

@router.get("/connections/connected")
def get_connected_users(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(MatchRequest).filter(
        MatchRequest.status == "accepted",
        or_(
            MatchRequest.sender_id == current_user.id,
            MatchRequest.receiver_id == current_user.id,
        )
    )

    total = query.count()
    skip = (page - 1) * limit
    items = query.offset(skip).limit(limit).all()

    users = []
    for req in items:
        if req.sender_id == current_user.id:
            users.append(req.receiver)
        else:
            users.append(req.sender)

    has_next = skip + limit < total
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "has_next": has_next,
        "users": users
    }



@router.get("/recommended")
def get_recommended_users(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    blocked_ids = db.query(BlockedProfile.blocked_user_id).filter(
        BlockedProfile.blocker_id == current_user.id
    )
    matched_ids = db.query(MatchRequest.receiver_id).filter(
        MatchRequest.sender_id == current_user.id
    )
    saved_ids = db.query(SavedProfile.saved_user_id).filter(
        SavedProfile.user_id == current_user.id
    )

    query = db.query(User).filter(
        User.id != current_user.id,
        ~User.id.in_(blocked_ids),
        ~User.id.in_(matched_ids),
        ~User.id.in_(saved_ids),
        User.status.in_(allowed_statuses)
    ).order_by(func.random())  # Random ordering

    total = query.count()
    skip = (page - 1) * limit
    users = query.offset(skip).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "has_next": skip + limit < total,
        "users": users
    }


@router.get("/nearby")
def get_nearby_users(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    blocked_ids = db.query(BlockedProfile.blocked_user_id).filter(
        BlockedProfile.blocker_id == current_user.id
    )
    matched_ids = db.query(MatchRequest.receiver_id).filter(
        MatchRequest.sender_id == current_user.id
    )
    saved_ids = db.query(SavedProfile.saved_user_id).filter(
        SavedProfile.user_id == current_user.id
    )

    postal_code_prefix = current_user.postal[-3:] if current_user.postal and len(current_user.postal) >= 3 else ""

    query = db.query(User).filter(
        User.id != current_user.id,
        ~User.id.in_(blocked_ids),
        ~User.id.in_(matched_ids),
        ~User.id.in_(saved_ids),
        User.status.in_(allowed_statuses),
        or_(
            User.city_name == current_user.city_name,
            User.state == current_user.state,
            User.postal.like(f"%{postal_code_prefix}")
        )
    ).order_by(desc(User.lastSeen))

    total = query.count()
    skip = (page - 1) * limit
    users = query.offset(skip).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "has_next": skip + limit < total,
        "users": users
    }


@router.get("/new")
def get_new_profiles(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    blocked_ids = db.query(BlockedProfile.blocked_user_id).filter(
        BlockedProfile.blocker_id == current_user.id
    )
    matched_ids = db.query(MatchRequest.receiver_id).filter(
        MatchRequest.sender_id == current_user.id
    )
    saved_ids = db.query(SavedProfile.saved_user_id).filter(
        SavedProfile.user_id == current_user.id
    )

    query = db.query(User).filter(
        User.id != current_user.id,
        ~User.id.in_(blocked_ids),
        ~User.id.in_(matched_ids),
        ~User.id.in_(saved_ids),
        User.status.in_(allowed_statuses)
    ).order_by(desc(User.created_at))   # ✅ order by signup date

    total = query.count()
    skip = (page - 1) * limit
    users = query.offset(skip).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "has_next": skip + limit < total,
        "users": users
    }





@router.post("/edit-profile")
async def edit_profile(
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(...),
    bio: str = Form(""),
    occupation: str = Form(""),
    education: str = Form(""),
    city_name: str = Form(""),
    country_name: str = Form(""),
    pin_code: str = Form(""),
    partnerExpectations: str = Form(""),
    phonehide: bool = Form(False),
    photoProtect: bool = Form(False),
    photo1: Optional[UploadFile] = File(None),
    photo2: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user = db.query(User).filter(User.id == current_user.id).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    os.makedirs(UPLOAD_DIR, exist_ok=True)

    # Handle photo1
    if photo1:
        # Remove old photo1 if exists
        if user.photo1:
            old_path = os.path.join(UPLOAD_DIR, user.photo1)
            try:
                os.remove(old_path)
            except FileNotFoundError:
                pass

        # Save new photo1
        photo1_filename = f"{user.id}_photo1_{photo1.filename}"
        photo1_path = os.path.join(UPLOAD_DIR, photo1_filename)
        with open(photo1_path, "wb") as f:
            shutil.copyfileobj(photo1.file, f)
        user.photo1 = photo1_filename

        # mark approval false when photo1 changed
        # make sure your User model has attribute 'photo1Approve'
        user.photo1Approve = False

    # Handle photo2
    if photo2:
        # Remove old photo2 if exists
        if user.photo2:
            old_path = os.path.join(UPLOAD_DIR, user.photo2)
            try:
                os.remove(old_path)
            except FileNotFoundError:
                pass

        # Save new photo2
        photo2_filename = f"{user.id}_photo2_{photo2.filename}"
        photo2_path = os.path.join(UPLOAD_DIR, photo2_filename)
        with open(photo2_path, "wb") as f:
            shutil.copyfileobj(photo2.file, f)
        user.photo2 = photo2_filename

        # mark approval false when photo2 changed
        # make sure your User model has attribute 'photo2Approve'
        user.photo2Approve = False

    # Update other profile fields
    user.name = name
    user.email = email
    user.phone = phone

    user.occupation = occupation
    user.education = education
    user.city_name = city_name

    user.phonehide = phonehide
    user.photoProtect = photoProtect
    user.country = country_name
    user.postal = pin_code

    # partnerExpectations: set approval false only when changed
    # normalize comparison a bit (strip) to avoid trivial whitespace changes triggering approval reset
    if (partnerExpectations or "") != (user.partnerExpectations or ""):
        user.partnerExpectations = partnerExpectations
        # ensure attribute exists on model
        user.partnerExpectations_approval = False
    else:
        # no change -> keep existing approval
        user.partnerExpectations = partnerExpectations

    # bio: set approval false only when changed
    if (bio or "") != (user.bio or ""):
        user.bio = bio
        user.bio_approval = False
    else:
        user.bio = bio

    db.commit()
    db.refresh(user)

    return {"success": True, "message": "Profile updated successfully", "user": user}



@router.get("/my-views")
def get_my_views(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    views = db.query(ProfileView).filter(ProfileView.user_id == current_user.id).order_by(ProfileView.viewed_at.desc()).all()
    return views

@router.get("/self-views")
def get_self_views(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    views = db.query(User).filter(User.id == current_user.id).first()
    return views


from sqlalchemy.orm import joinedload
#from typing import Optional, Union
# @router.get("/notification/{status}")
# def get_user_notifications(
#     status: str,
#     db: Session = Depends(get_db),
#     current_user: User = Depends(get_current_user)
# ):
#     """message, view,'pending', 'accepted', 'rejected'"""
#     if status == "message":
#         notifications = (
#             db.query(Notification)
#             .options(joinedload(Notification.sender))
#             .filter(Notification.status == NotificationStatus.msg)
#             .order_by(Notification.created_at.desc())
#             .limit(25)
#             .all()
#         )
#     else:
#         notifications = (
#             db.query(Notification)
#             .options(joinedload(Notification.sender))
#             .filter(
#                 Notification.receiver_id == current_user.id,
#                 Notification.status == NotificationStatus(status)
#             )
#             .order_by(Notification.created_at.desc())
#             .limit(25)
#             .all()
#         )
#
#     result = []
#
#     # ✅ Mark notifications as read after fetching
#     for n in notifications:
#         if not n.is_read:   # only update if not already read
#             n.is_read = True
#
#         sender_data = None
#         if n.sender:
#             sender_data = n.sender
#
#         result.append({
#             "id": n.id,
#             "sender_id": sender_data,
#             "receiver_id": n.receiver_id,
#             "status": n.status,
#             "message": n.message or "",
#             "is_read": n.is_read,
#             "created_at": n.created_at.isoformat() if n.created_at else None
#         })
#
#     # ✅ Persist the read status in DB
#     db.commit()
#
#     return result



@router.get("/notification/{status}")
def get_user_notifications(
    status: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """message, view,'pending', 'accepted', 'rejected'"""
    if status == "message":
        notifications = (
            db.query(Notification)
            .options(joinedload(Notification.sender))
            .filter(Notification.status == NotificationStatus.msg)
            .order_by(Notification.created_at.desc())
            .limit(25)
            .all()
        )
    else:
        notifications = (
            db.query(Notification)
            .options(joinedload(Notification.sender))
            .filter(
                Notification.receiver_id == current_user.id,
                Notification.status == NotificationStatus(status)
            )
            .order_by(Notification.created_at.desc())
            .limit(25)
            .all()
        )

    result = []
    updated = False

    for n in notifications:
        if not n.is_read:   # only update if not already read
            n.is_read = True
            updated = True
        sender_data = None
        if n.sender:
            sender_data=n.sender

        result.append({
            "id": n.id,
            "sender_id": sender_data,
            "receiver_id": n.receiver_id,
            "status": n.status,
            "message": n.message or "",
            "is_read": n.is_read,
            "created_at": n.created_at.isoformat() if n.created_at else None
        })
    if updated:
        db.commit()
    return result


@router.get("/notification_get/unread-count")
def get_unread_notification_counts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Return unread notification counts grouped by type and total"""

    # Base query for unread notifications for current user
    base_query = (
        db.query(Notification.status, func.count(Notification.id))
        .filter(
            Notification.receiver_id == current_user.id,
            Notification.is_read == False
        )
        .group_by(Notification.status)
        .all()
    )

    # Build result dict with default 0 counts
    counts = {
        "msg": 0,
        "view": 0,
        "pending": 0,
        "accepted": 0,
        "rejected": 0,
    }

    # Map DB values directly
    # for status, count in base_query:
    #     counts[status] = count  # <-- use status directly
    #
    # # Add total
    # counts["total_unread"] = sum(counts.values())
    #
    # return counts
    for status, count in base_query:
        key = status.value if hasattr(status, "value") else str(status)
        if key in counts:
            counts[key] = count

    counts["total_unread"] = sum(counts.values())
    return counts


@router.get("/users/list", response_model=List[UserPublicResponse])
def get_users_list( current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    users = db.query(User).filter(User.id != current_user.id).all()

    response = []
    for user in users:
        public_data = get_user_public_response(user, current_user.id, db)
        response.append(public_data)

    return response





@router.get("/view_profile/{userId}", response_model=UserPublicResponse)
def get_user_profile(
        userId:str,
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db),
):
    create_or_update_view_notification(
        sender_id=current_user.id,
        receiver_id=userId,
        db=db
    )
    return build_user_response(userId, current_user.id, db)



def calculate_age(dob: date) -> int:
    today = date.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

def compute_age(dob: Optional[date]) -> Optional[int]:
    if not dob:
        return None
    today = date.today()
    age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
    return age


@router.get("/profiles", response_model=ProfilesResponse)
async def get_profiles(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    user_id: Optional[str] = Query(None),
    min_age: Optional[int] = Query(18),
    max_age: Optional[int] = Query(65),
    location: Optional[str] = Query(None),
    genders: Optional[str] = Query(None),
    limit: int = Query(10),
    offset: int = Query(0)
):
    # clamp pagination
    limit = max(1, min(limit, 100))
    offset = max(0, offset)

    # statuses to exclude
    exclude_statuses = ["deleted", "deleted_by_user", "banned", "suspended", "inactive"]

    # If caller asked for exact user_id, try exact lookup first (bypass signup check).
    # But still respect excluded statuses and exclude the current user if you want to.
    if user_id:
        u = db.query(User).filter(cast(User.id, String) == str(user_id)).first()
        if u:
            if u.status and u.status.lower() in [s.lower() for s in exclude_statuses]:
                # explicitly excluded by status
                return ProfilesResponse(total=0, offset=offset, limit=limit, users=[])
            if str(u.id) == str(current_user.id):
                # requested user is current user: keep same behavior as list (excluded)
                return ProfilesResponse(total=0, offset=offset, limit=limit, users=[])
            # Optionally check blocked/saved/match flags for this single user:
            is_saved = db.query(SavedProfile).filter(
                SavedProfile.user_id == current_user.id,
                SavedProfile.saved_user_id == u.id
            ).first() is not None

            is_blocked = db.query(BlockedProfile).filter(
                BlockedProfile.blocker_id == current_user.id,
                BlockedProfile.blocked_user_id == u.id
            ).first() is not None

            is_blocked_by_self = db.query(BlockedProfile).filter(
                BlockedProfile.blocker_id == u.id,
                BlockedProfile.blocked_user_id == current_user.id
            ).first() is not None

            m_request = db.query(MatchRequest).filter(
                ((MatchRequest.sender_id == current_user.id) & (MatchRequest.receiver_id == u.id)) |
                ((MatchRequest.sender_id == u.id) & (MatchRequest.receiver_id == current_user.id))
            ).order_by(MatchRequest.created_at.desc()).first()

            is_matched = False
            match_status = "none"
            if m_request:
                match_status = getattr(m_request, "status", "pending")
                is_matched = match_status == "accepted"

            u_dict = UserPublicResponse.from_attributes(u).model_dump() if hasattr(UserPublicResponse, "from_attributes") else UserPublicResponse.from_orm(u).dict()
            u_dict["age"] = compute_age(u.dob)
            u_dict["isSaved"] = is_saved
            u_dict["isBlocked"] = is_blocked
            u_dict["isBlockedBySelf"] = is_blocked_by_self
            u_dict["isMatched"] = is_matched
            u_dict["match_status"] = match_status

            return ProfilesResponse(total=1, offset=offset, limit=limit, users=[UserPublicResponse(**u_dict)])
        else:
            # user_id not found — return empty result
            return ProfilesResponse(total=0, offset=offset, limit=limit, users=[])

    # Build base query — NOTE: removed is_signup_complete filter per your request.
    query = db.query(User).filter(
        cast(User.id, String) != str(current_user.id),
        func.lower(User.status).notin_([s.lower() for s in exclude_statuses])
    )

    # Age range filter (min_age/min_age)
    if min_age is not None or max_age is not None:
        min_age = min_age if min_age is not None else 18
        max_age = max_age if max_age is not None else 100
        today = date.today()
        earliest_dob = date(today.year - max_age, today.month, today.day)
        latest_dob = date(today.year - min_age, today.month, today.day)
        query = query.filter(User.dob.between(earliest_dob, latest_dob))

    # Location filter
    if location:
        loc = f"%{location.lower()}%"
        query = query.filter(
            func.lower(User.city_name).ilike(loc) |
            func.lower(User.state).ilike(loc) |
            func.lower(User.country).ilike(loc)
        )

    # Gender filter
    if genders:
        gender_list = [g.strip().lower() for g in genders.split(",") if g.strip()]
        if gender_list:
            query = query.filter(func.lower(User.gender).in_(gender_list))

    # Exclude users that current_user blocked/matched/saved (use subqueries)
    blocked_subq = db.query(BlockedProfile.blocked_user_id).filter(
        BlockedProfile.blocker_id == current_user.id
    ).subquery()

    matched_subq = db.query(MatchRequest.receiver_id).filter(
        MatchRequest.sender_id == current_user.id
    ).subquery()

    saved_subq = db.query(SavedProfile.saved_user_id).filter(
        SavedProfile.user_id == current_user.id
    ).subquery()

    query = query.filter(
        ~cast(User.id, String).in_(blocked_subq),
        ~cast(User.id, String).in_(matched_subq),
        ~cast(User.id, String).in_(saved_subq)
    )

    # Ordering (MySQL-compatible; avoid NULLS LAST)
    if hasattr(User, "lastSeen"):
        query = query.order_by((User.lastSeen == None), User.lastSeen.desc())
    elif hasattr(User, "created_at"):
        query = query.order_by((User.created_at == None), User.created_at.desc())
    else:
        query = query.order_by(User.id)

    total = query.count()
    users = query.offset(offset).limit(limit).all()

    # Batch fetch relationships to avoid N+1 queries:
    user_ids = [str(u.id) for u in users]
    saved_map = {}
    blocked_map = {}
    blocked_by_self_map = {}
    match_map: Dict[str, Dict[str, Any]] = {}  # user_id -> {"isMatched": bool, "match_status": str}

    if user_ids:
        # Saved (does current_user saved these users?)
        saved_rows = db.query(SavedProfile.saved_user_id).filter(
            SavedProfile.user_id == current_user.id,
            SavedProfile.saved_user_id.in_(user_ids)
        ).all()
        saved_set = {r[0] for r in saved_rows}

        # Blocked by current_user
        blocked_rows = db.query(BlockedProfile.blocked_user_id).filter(
            BlockedProfile.blocker_id == current_user.id,
            BlockedProfile.blocked_user_id.in_(user_ids)
        ).all()
        blocked_set = {r[0] for r in blocked_rows}

        # Blocked by those users (did they block current_user?)
        blocked_by_self_rows = db.query(BlockedProfile.blocker_id).filter(
            BlockedProfile.blocker_id.in_(user_ids),
            BlockedProfile.blocked_user_id == current_user.id
        ).all()
        blocked_by_self_set = {r[0] for r in blocked_by_self_rows}

        # MatchRequests (either direction) between current_user and each user_id
        match_rows = db.query(MatchRequest.sender_id, MatchRequest.receiver_id, MatchRequest.status).filter(
            ((MatchRequest.sender_id == current_user.id) & (MatchRequest.receiver_id.in_(user_ids))) |
            ((MatchRequest.receiver_id == current_user.id) & (MatchRequest.sender_id.in_(user_ids)))
        ).all()
        # Build map keyed by the other user's id
        for sender_id, receiver_id, status in match_rows:
            other_id = receiver_id if sender_id == current_user.id else sender_id
            is_matched = (status == "accepted")
            match_map[str(other_id)] = {"isMatched": is_matched, "match_status": status or "pending"}

        # assign sets to maps for quick lookup
        saved_map = {uid: (uid in saved_set) for uid in user_ids}
        blocked_map = {uid: (uid in blocked_set) for uid in user_ids}
        blocked_by_self_map = {uid: (uid in blocked_by_self_set) for uid in user_ids}
    else:
        saved_map = {}
        blocked_map = {}
        blocked_by_self_map = {}
        match_map = {}

    # Build response users
    users_out: List[UserPublicResponse] = []
    for u in users:
        uid = str(u.id)
        # serialize with pydantic
        if hasattr(UserPublicResponse, "from_attributes"):
            base = UserPublicResponse.from_attributes(u).model_dump()
        else:
            base = UserPublicResponse.from_orm(u).dict()

        base["age"] = compute_age(u.dob)
        base["isSaved"] = saved_map.get(uid, False)
        base["isBlocked"] = blocked_map.get(uid, False)
        base["isBlockedBySelf"] = blocked_by_self_map.get(uid, False)
        match_info = match_map.get(uid, {"isMatched": False, "match_status": "none"})
        base["isMatched"] = match_info["isMatched"]
        base["match_status"] = match_info["match_status"]

        users_out.append(UserPublicResponse(**base))

    return ProfilesResponse(total=total, offset=offset, limit=limit, users=users_out)



class UserDataSchema(BaseModel):
    userId: str
    device_id: str


@router.post("/api/receive_user_data")
def receive_user_data(data: UserDataSchema, db: Session = Depends(get_db)):
    try:
        user_token = data.userId
        device_id = data.device_id

        # Check if user exists
        user = db.query(User).filter_by(id=user_token).first()

        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        # Update the device token
        user.devicetoken = device_id
        db.commit()

        return {"message": "Data received successfully"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")

def generate_order_id():
    # Generate a unique identifier using uuid
    unique_id = str(uuid.uuid4())

    # Extract the first 7 characters to create a 7-digit order ID
    order_id = unique_id[:7]

    return order_id

@router.post("/app_pay")
async def api_app_pay(amount:int, currency:str,current_user: User = Depends(get_current_user),  db: Session = Depends(get_db)):
    try:

        currency = currency
        amount = amount

        if not currency or not amount:
            raise HTTPException(status_code=400, detail="Currency and amount are required.")

        order_id = generate_order_id()
        response_data = {
            "MatriID": current_user.id,
            "currency": currency,
            "amount": amount,
            "order_id": order_id,
            "email":current_user.email,
            "contact":current_user.mobile
        }


        return JSONResponse(content=response_data, status_code=200)

    except Exception as e:
        #print(f"Error: {str(e)}")
        return JSONResponse(content={"status": "error"}, status_code=500)


class PaymentRequest(BaseModel):
    userId: str
    currency: str
    amount: int
    membershiptype:str

@router.post("/api/payment_success")
async def payment_success(data: PaymentRequest, db: Session = Depends(get_db)):
    try:
        today = datetime.today()
        id = data.userId
        currency = data.currency
        amount = data.amount
        memtype = data.membershiptype
        if not id or not currency or not amount:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Missing userId, currency, or amount in the request"
            )

        # Get user by email
        user = db.query(User).filter(User.id == id).first()

        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        # Create new payment object
        new_payment = Payment(
            user_id=user.id,
            mobile_no=user.mobile if user.mobile else None,
            email_id=user.email,
            currency=currency,
            amount=int(amount),
            order_id="Razor",
            payment_id="Razor",
            status="Success",
            date=today
        )

        # Add and commit new payment
        db.add(new_payment)
        db.commit()
        plan = db.query(MembershipPlan).filter(MembershipPlan.membership_name == memtype).first()
        if not plan:
            raise HTTPException(status_code=404, detail="Membership not found")
        # Update user subscription details based on currency and amount

        user.Status = plan.status
        user.video_min =plan.video_mins
        user.voice_min = plan.voice_mins
        user.memtype=memtype
        user.membershipExpiryDate = date.today() + timedelta(days=plan.days)

        db.commit()

        return JSONResponse(
            content={"message": "Data received successfully"},
            status_code=status.HTTP_200_OK
        )

    except Exception as e:
        return JSONResponse(
            content={"error": f"An error occurred: {str(e)}"},
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@router.post("/send-notification")
def send_notification(data: NotificationRequest):
    success = send_notification_to_all_user(
        title=data.title,
        body=data.body,
        image=data.image,
        token=data.token
    )

    if success:
        return {"status": "success", "message": "Notification sent successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to send notification")


# Temporary storage for OTPs (Use Redis or DB in production)
otp_store_recover = {}

class AccountRecoveryRequest(BaseModel):
    contact_info: str
    mobile_code:str

class OTPVerificationRequest(BaseModel):
    contact_info: str
    mobile_code: str
    otp: str
class PasswordRecoveryRequest(BaseModel):
    contact_info: str
    mobile_code:str

class PasswordOTPVerificationRequest(BaseModel):
    contact_info: str
    mobile_code: str
    password:str
    otp: str


# Step 1: Request OTP

otp_store = {}
@router.post("/api/recover_account", status_code=status.HTTP_200_OK)
async def recover_account(
    data: AccountRecoveryRequest,
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(
        or_(User.mobile == data.contact_info, User.email == data.contact_info)
    ).first()

    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    if user.status != "Deleted":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Account is already active")

    otp = str(random.randint(100000, 999999))

    # Delete any old OTPs for this user
    db.query(UserOTP).filter_by(user_id=user.id).delete()

    # Save new OTP
    db.add(UserOTP(user_id=user.id, otp=otp, created_at=datetime.now()))
    db.commit()
    full_number=data.mobile_code + data.contact_info
    sent = send_otp_sms(full_number, otp)
    if not sent:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to send OTP")

    masked_mobile = f"{user.mobile[:5]}xxxxx"
    return {
        "success": True,
        "message": f"OTP sent successfully to {masked_mobile}"
    }


# Step 2: Verify OTP and Restore Account
@router.post("/api/recover_account_verify", status_code=status.HTTP_200_OK)
async def recover_account_verify(
    data: OTPVerificationRequest,
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(
        or_(User.mobile == data.contact_info, User.email == data.contact_info)
    ).first()

    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    otp_entry = db.query(UserOTP).filter_by(user_id=user.id, otp=data.otp).first()
    if not otp_entry:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid OTP")

    try:
        created_at = otp_entry.created_at.replace(tzinfo=india_tz)
    except Exception:
        created_at = otp_entry.created_at  # fallback if already aware

    # Check expiry (30 minutes)
    now = datetime.now(india_tz)
    if now - created_at > timedelta(minutes=30):
        db.delete(otp_entry)
        db.commit()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="OTP expired")

    # Reactivate account
    user.status = "Active"
    db.delete(otp_entry)
    db.commit()

    return {
        "success": True,
        "message": "User account recovered successfully"
    }

@router.post("/api/forget-password", status_code=status.HTTP_200_OK)
async def forget_password(
    data: PasswordRecoveryRequest,
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(
        or_(User.mobile == data.contact_info, User.email == data.contact_info)
    ).first()

    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")


    otp = str(random.randint(100000, 999999))

    # Delete any old OTPs for this user
    db.query(UserOTP).filter_by(user_id=user.id).delete()

    # Save new OTP
    db.add(UserOTP(user_id=user.id, otp=otp, created_at=datetime.now()))
    db.commit()
    full_number=data.mobile_code + data.contact_info
    sent = send_otp_sms(full_number, otp)
    if not sent:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to send OTP")

    masked_mobile = f"{user.mobile[:5]}xxxxx"
    return {
        "success": True,
        "message": f"OTP sent successfully to {masked_mobile}"
    }


# Step 2: Verify OTP and Restore Account
@router.post("/api/forget_password_verify", status_code=status.HTTP_200_OK)
async def password_otp_verification_request(
    data: PasswordOTPVerificationRequest,
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(
        or_(User.mobile == data.contact_info, User.email == data.contact_info)
    ).first()

    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    otp_entry = db.query(UserOTP).filter_by(user_id=user.id, otp=data.otp).first()
    if not otp_entry:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid OTP")

    try:
        created_at = otp_entry.created_at.replace(tzinfo=india_tz)
    except Exception:
        created_at = otp_entry.created_at  # fallback if already aware

    # Check expiry (30 minutes)
    now = datetime.now(india_tz)
    if now - created_at > timedelta(minutes=30):
        db.delete(otp_entry)
        db.commit()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="OTP expired")

    # Reactivate account
    user.mobileverify=True
    user.password = set_password(data.password),
    db.delete(otp_entry)
    db.commit()

    return {
        "success": True,
        "message": "User account password recovered successfully"
    }


@router.post("/api/delete_user_profile", status_code=status.HTTP_200_OK)
async def delete_user_profile(

    current_user: User = Depends(get_current_user), # Extract token from Authorization header
    db: Session = Depends(get_db)
):
    #

    if not current_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    # Delete user profile (mark as deleted)
    current_user.status = "Deleted"
    db.commit()

    return {
        "success": True,
        "message": "User profile deleted successfully"
    }



@router.get("/api/world_mobile_codes")
async def world_mobile_codes():
    """API Endpoint to Get World Mobile Codes Without +."""
    codes = get_world_mobile_codes()
    return {"status": "success", "data": codes}

import io

def yesno_to_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ['yes', 'true', '1']
    if isinstance(value, int):
        return value == 1
    return False
def validate_length(val, max_len, field):
    if val and len(str(val)) > max_len:
        raise ValueError(f"{field} exceeds max length ({max_len})")
    return val

# Example usage:


@router.post("/upload-users-excel")
async def upload_users_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed.")

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents), data_only=True)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]
    expected_headers = [
        "id", "email", "password", "name", "gender", "dob", "age", "maritalStatus", "education", "occupation",
        "language", "height", "diet", "smoke", "drink", "city_name", "postal", "state", "country",
        "mobile", "phonehide", "mobilecode", "partnerExpectations", "bio", "status", "memtype",
        "membershipExpiryDate", "photoProtect", "chatcontact", "devicetoken", "pagecount", "onlineUsers",
        "mobileverify", "verify_status", "verify_email", "video_min", "voice_min", "photo1",
        "photo1Approve", "photo2", "photo2Approve", "chat_msg", "photohide", "lastSeen"
    ]

    if headers != expected_headers:
        raise HTTPException(status_code=400, detail="Excel headers do not match expected format.")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            user = User(
                id=row[0],
                email=row[1],
                password=row[2],
                name=row[3],
                gender=row[4],
                dob=datetime.strptime(row[5], "%Y-%m-%d").date() if row[5] else None,
                age=int(row[6]) if row[6] else None,
                maritalStatus=row[7],
                education=row[8],
                occupation=row[9],
                language=row[10],
                height=None,
                diet=row[12],
                smoke=row[13],
                drink=row[14],
                city_name=row[15],
                postal=row[16],
                #postal=validate_length(row[16], 20, "postal"),

                state=row[17],
                country=row[18],
                mobile=row[19],
                phonehide=yesno_to_bool(row[20]),
                mobilecode=row[21],
                partnerExpectations=row[22],
                bio=row[23],
                status=row[24],
                memtype=row[25],
                membershipExpiryDate=datetime.strptime(row[26], "%Y-%m-%d").date() if row[26] else None,
                photoProtect=yesno_to_bool(row[27]),
                chatcontact=yesno_to_bool(row[28]),
                devicetoken=row[29],
                pagecount=int(row[30]) if row[30] else 0,
                onlineUsers=yesno_to_bool(row[31]),
                mobileverify=yesno_to_bool(row[32]),
                verify_status=yesno_to_bool(row[33]),
                verify_email=yesno_to_bool(row[34]),
                video_min=int(row[35]) if row[35] else 0,
                voice_min=int(row[36]) if row[36] else 0,
                photo1=row[37],
                photo1Approve=yesno_to_bool(row[38]),
                photo2=row[39],
                photo2Approve=yesno_to_bool(row[40]),
                chat_msg=int(row[41]) if row[41] else 0,
                photohide=yesno_to_bool(row[42]),
                lastSeen=datetime.strptime(row[43], "%Y-%m-%d %H:%M:%S") if row[43] else None
            )

            db.add(user)

        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error in row {row[0]}: {str(e)}")

    db.commit()
    return {"message": "Users uploaded successfully."}

class ContactCreate(BaseModel):
    subject: str = Field(..., min_length=3)
    message: str = Field(..., min_length=10)

def create_contact(db: Session, contact_data: ContactCreate, user: User):
    contact = Contact(
        subject=contact_data.subject,
        message=contact_data.message,
        user_id=user.id
    )
    db.add(contact)
    db.commit()
    db.refresh(contact)
    return contact

@router.post("/contact", status_code=status.HTTP_201_CREATED)
def submit_contact_form(
    contact: ContactCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return create_contact(db, contact, current_user)

@router.post("/chat/check-send/{receiver_id}")
def check_and_send_message(
    receiver_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
     # Import your function

    success = check_and_log_chat(db, receiver_id, current_user)
    return {"success": success}

@router.post("/requests/withdrawal")
def withdrawal_match_request(
    payload: RequestPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    receiver_id = payload.user_id

    # ✅ 1. Prevent self withdrawal (must target another user)
    if current_user.id == receiver_id:
        raise HTTPException(status_code=400, detail="Invalid withdrawal request.")

    # ✅ 2. Ensure receiver exists
    receiver = db.query(User).filter(User.id == receiver_id).first()
    if not receiver:
        raise HTTPException(status_code=404, detail="Receiver user not found.")

    # ✅ 3. Find existing match request
    existing_request = db.query(MatchRequest).filter_by(
        sender_id=current_user.id,
        receiver_id=receiver_id
    ).first()

    if not existing_request:
        raise HTTPException(status_code=404, detail="No active match request to withdraw.")

    # ✅ 4. Delete match request
    db.delete(existing_request)
    db.commit()

    # ✅ 5. Optionally: delete related notification
    existing_notify = db.query(Notification).filter_by(
        sender_id=current_user.id,
        receiver_id=receiver_id,
        status=NotificationStatus.pending
    ).first()
    if existing_notify:
        db.delete(existing_notify)
        db.commit()

    return {
        "success": True,
        "message": "Match request withdrawn successfully."
    }



RAZOR_KEY_ID = os.environ.get("rzp_live_Z0qfc9VQ6G85BW")
RAZOR_KEY_SECRET = os.environ.get("xUNObYwF43SOr1sDioxV7xioT")
client = razorpay.Client(auth=("rzp_live_Z0qfc9VQ6G85BW", "xUNObYwF43SOr1sDioxV7xio"))
#client = razorpay.Client(auth=(RAZOR_KEY_ID, RAZOR_KEY_SECRET))

class CreateOrderReq(BaseModel):
    userId: str = Field(..., alias="userId")
    amount: int   # amount in subunits (paise for INR, cents for USD)
    currency: str
    receipt: str
    plan_id: str | None = None


@router.post("/create-order")
def create_order(payload: CreateOrderReq, db: Session = Depends(get_db)):
    # 1) Basic checks
    user = db.query(User).filter(User.id == payload.userId).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # 2) Get plan if provided
    plan = None
    if payload.plan_id:
        plan = db.query(MembershipPlan).filter(MembershipPlan.membership_name == payload.plan_id).first()
        if not plan:
            raise HTTPException(status_code=404, detail="Membership plan not found")

        # 3) Determine expected amount (in subunits). Prefer plan fields if present
        # Expect that MembershipPlan has price fields: price_in_inr and price_usd (or adapt)
        if payload.currency.upper() == "INR":
            if getattr(plan, "inr_price", None) is not None:
                expected_subunits = int(round(plan.inr_price))
            else:
                # fallback to `price` field assumed to be INR value
                expected_subunits = int(round((getattr(plan, "price", 0)) * 100))
        else:  # USD or other currencies (we only support USD here)
            if getattr(plan, "usd_price", None) is not None:
                expected_subunits = int(round(plan.usd_price ))
            else:
                # fallback: optionally keep a mapping or deny if unknown
                raise HTTPException(status_code=400, detail="Plan price for this currency is not configured")

        if payload.amount != expected_subunits:
            raise HTTPException(
                status_code=400,
                detail=f"Amount mismatch. Expected {expected_subunits} (subunits) for plan {plan.membership_name}"
            )

    # 4) Create Razorpay order
    try:
        order_payload = {
            "amount": payload.amount,
            "currency": payload.currency.upper(),
            "receipt": payload.receipt,
            "payment_capture": 1
        }
        order = client.order.create(order_payload)

        # 5) Persist an order/payment row with status 'created' (adapt Payment model fields)
        new_payment = Payment(
            user_id=user.id,
            email_id=user.email,
            mobile_no=user.mobile if hasattr(user, "mobile") else None,
            currency=order.get("currency"),
            amount=order.get("amount"),
            order_id=order.get("id"),          # store razorpay order id
            payment_id=None,                   # will be set on verification
            status="Initiated",
            date=datetime.now(india_tz),
            #plan_id=payload.plan_id
        )
        db.add(new_payment)
        db.commit()

        return {
            "id": order["id"],
            "amount": order["amount"],
            "currency": order["currency"],
            "order": order
        }
    except Exception as e:
        # you may want to log e
        raise HTTPException(status_code=500, detail=str(e))


class VerifyReq(BaseModel):
    userId: str = Field(..., alias="userId")
    plan_id: str | None = None
    amount: int | None = None
    currency: str | None = None
    razorpay_payment_id: str
    razorpay_order_id: str
    razorpay_signature: str


@router.post("/verify")
def verify_payment(payload: VerifyReq, db: Session = Depends(get_db)):
    # Verify signature using Razorpay utility
    client_local = razorpay.Client(auth=(RAZOR_KEY_ID, RAZOR_KEY_SECRET))
    try:
        params = {
            "razorpay_order_id": payload.razorpay_order_id,
            "razorpay_payment_id": payload.razorpay_payment_id,
            "razorpay_signature": payload.razorpay_signature
        }
        client_local.utility.verify_payment_signature(params)
    except razorpay.errors.SignatureVerificationError:
        raise HTTPException(status_code=400, detail="Invalid signature")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Razorpay verification error: {str(e)}")

    # Begin DB transaction: update payment record and user's membership atomically
    try:
        # find the created Payment record by order_id
        payment_record = db.query(Payment).filter(Payment.order_id == payload.razorpay_order_id).first()

        if not payment_record:
            # fallback: maybe the client didn't call create-order; we still create a record
            payment_record = Payment(
                user_id=payload.userId,
                email_id=None,
                mobile_no=None,
                currency=payload.currency,
                amount=payload.amount,
                order_id=payload.razorpay_order_id,
                payment_id=payload.razorpay_payment_id,
                status="success",
                created_at=datetime.utcnow(),
                plan_id=payload.plan_id
            )
            db.add(payment_record)
            db.commit()  # commit this quick fallback
        else:
            # update existing record
            payment_record.payment_id = payload.razorpay_payment_id
            payment_record.status = "success"
            payment_record.signature = payload.razorpay_signature if hasattr(payment_record, "signature") else None
            payment_record.updated_at = datetime.utcnow()
            db.commit()

        # Load user & plan and update membership details
        user = db.query(User).filter(User.id == payload.userId).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        plan = None
        if payload.plan_id:
            plan = db.query(MembershipPlan).filter(MembershipPlan.membership_name == payload.plan_id).first()
            if not plan:
                raise HTTPException(status_code=404, detail="Membership plan not found")

        # (Optional) validate that the paid amount matches expected server-side amount
        if plan and payload.amount is not None and payload.currency:
            if payload.currency.upper() == "INR":
                expected = int(round(getattr(plan, "price_in_inr", getattr(plan, "price", 0)) * 100))
            else:
                expected = int(round(getattr(plan, "price_usd", None) * 100)) if getattr(plan, "price_usd", None) is not None else None

            if expected is not None and payload.amount != expected:
                raise HTTPException(status_code=400, detail="Paid amount does not match expected plan price")

        # Now update user membership (atomically)
        # Use today's date or existing expiry (whichever is later) so upgrade extends appropriately
        today = date.today()
        current_expiry = getattr(user, "membershipExpiryDate", None)
        start_date = today
        if current_expiry and isinstance(current_expiry, date) and current_expiry > today:
            start_date = current_expiry

        if plan:
            days = int(getattr(plan, "days", 0) or 0)
            new_expiry = start_date + timedelta(days=days)

            user.status = getattr(plan, "status", getattr(user, "status", "active"))
            user.video_min = getattr(plan, "video_mins", getattr(user, "video_min", user.video_min))
            user.voice_min = getattr(plan, "voice_mins", getattr(user, "voice_min", user.voice_min))
            user.memtype = plan.membership_name
            user.membershipExpiryDate = new_expiry
            # mark verified or active if not already
            user.verify_status = True
            user.mobileverify = True

            # persist changes
            db.commit()
        else:
            # No plan provided — don't change membership, but keep payment record
            pass

        return {
            "status": "success",
            "message": "Payment verified and membership updated",
            "membershipExpiryDate": str(user.membershipExpiryDate)
        }

    except HTTPException:
        # re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/payment_success")
async def payment_success(data: PaymentRequest, db: Session = Depends(get_db)):
    try:
        today = datetime.today()
        id = data.userId
        currency = data.currency
        amount = data.amount
        memtype = data.membershiptype
        if not id or not currency or not amount:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Missing userId, currency, or amount in the request"
            )

        # Get user by email
        user = db.query(User).filter(User.id == id).first()

        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        # Create new payment object
        new_payment = Payment(
            user_id=user.id,
            mobile_no=user.mobile if user.mobile else None,
            email_id=user.email,
            currency=currency,
            amount=int(amount),
            order_id="Razor",
            payment_id="Razor",
            status="Success",
            date=today
        )

        # Add and commit new payment
        db.add(new_payment)
        db.commit()
        plan = db.query(MembershipPlan).filter(MembershipPlan.membership_name == memtype).first()
        if not plan:
            raise HTTPException(status_code=404, detail="Membership not found")
        # Update user subscription details based on currency and amount

        user.Status = plan.status
        user.video_min =plan.video_mins
        user.voice_min = plan.voice_mins
        user.memtype=memtype
        user.membershipExpiryDate = date.today() + timedelta(days=plan.days)

        db.commit()

        return JSONResponse(
            content={"message": "Data received successfully"},
            status_code=status.HTTP_200_OK
        )

    except Exception as e:
        return JSONResponse(
            content={"error": f"An error occurred: {str(e)}"},
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@router.get("/users/online-offiline", status_code=200)
def get_online_users(
    status_online_offile: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    current_user.onlineUsers =  status_online_offile
    db.commit()
    db.refresh(current_user)

    return {"success": True, "message": "Online users fetched successfully", "user": current_user}
